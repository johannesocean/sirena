# Copyright (c) 2020 SMHI, Swedish Meteorological and Hydrological Institute.
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).
"""
Created on 2021-03-29 13:20

@author: johannes
"""
import shutil
import openpyxl
from openpyxl import load_workbook
from sirena.writers.writer import WriterBase


class ExcelTemplateWriter(WriterBase):
    """Template writer."""

    def __init__(self, template_path=None, export_path=None, cell_mapping=None, **kwargs):
        """Initialize."""
        super(ExcelTemplateWriter, self).__init__()
        self.workbook_template_path = template_path
        self.export_path = export_path
        self.cell_mapping = cell_mapping
        self.template_sheetname = 'mwreg'
        self.workbook = None
        for name, item in kwargs.items():
            setattr(self, name, item)

        self.copy_new_workbook()

    def write(self, data):
        """Write data to excel template."""
        ws = self.workbook[self.template_sheetname]

        row_number = self.start_row_index or 0
        for item in data.values():
            for attr, value in item.items():
                cell_id = self.cell_mapping.get(attr) % row_number
                ws[cell_id] = value
            row_number += 1
            if row_number == 50:
                row_number = 60

        if hasattr(self, 'img_path'):
            img = openpyxl.drawing.image.Image(self.img_path)
            img.anchor = 'B2'
            ws.add_image(img)

        self.workbook.template = False
        self.workbook.save(self.export_path)

    def copy_new_workbook(self):
        """Copy template workbook."""
        shutil.copyfile(self.workbook_template_path, self.export_path)
        self.workbook = load_workbook(self.export_path)

    def get_new_sheet(self, new_sheet_name):
        """Return new sheet."""
        ws = self.workbook.copy_worksheet(self.workbook[self.template_sheetname])
        ws.title = new_sheet_name
        return ws
